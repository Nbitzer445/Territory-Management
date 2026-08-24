"""Parsers for the three source spreadsheets, with de-dup / merge logic so
re-importing a fresh monthly export never overwrites manually-added
contacts, notes, or follow-ups, and sales history accumulates as snapshots
instead of being overwritten.
"""
import hashlib
import json
import re
from datetime import datetime, date

import openpyxl

MARKETS = ["Lincoln", "Omaha", "Norfolk", "Columbus", "Fremont", "Sioux City", "NW Iowa", "Other"]

_MARKET_KEYWORDS = [
    ("sioux center", "NW Iowa"),
    ("le mars", "NW Iowa"),
    ("spencer", "NW Iowa"),
    ("storm lake", "NW Iowa"),
    ("carroll", "NW Iowa"),
    ("south sioux city", "Sioux City"),
    ("sioux city", "Sioux City"),
    ("norfolk", "Norfolk"),
    ("columbus", "Columbus"),
    ("fremont", "Fremont"),
    ("council bluffs", "Omaha"),
    ("la vista", "Omaha"),
    ("lincoln", "Lincoln"),
    ("omaha", "Omaha"),
]


def infer_market(name, city=None):
    text = (city or name or "").lower()
    for kw, market in _MARKET_KEYWORDS:
        if kw in text:
            return market
    return "Other"


def _clean(v):
    if v is None:
        return ""
    return str(v).strip()


def _parse_date(v):
    """Accepts 'MM-DD-YYYY' strings, datetime/date objects, or None. Returns ISO 'YYYY-MM-DD' or None."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%m-%d-%Y", "%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _split_attendees(s):
    if not s:
        return []
    parts = re.split(r"[;,/]| and ", s)
    return [p.strip() for p in parts if p.strip()]


# ---------------------------------------------------------------------------
# Account / brand / contact upsert helpers (shared across importers + app)
# ---------------------------------------------------------------------------

def upsert_account(conn, name, **fields):
    name = _clean(name)
    if not name:
        return None
    row = conn.execute("SELECT * FROM accounts WHERE name = ? COLLATE NOCASE", (name,)).fetchone()
    if row is None:
        cols = ["name"]
        vals = [name]
        for k, v in fields.items():
            if v not in (None, ""):
                cols.append(k)
                vals.append(v)
        placeholders = ",".join("?" for _ in cols)
        conn.execute(f"INSERT INTO accounts ({','.join(cols)}) VALUES ({placeholders})", vals)
        cur = conn.execute("SELECT id FROM accounts WHERE name = ? COLLATE NOCASE", (name,))
        return cur.fetchone()["id"]
    else:
        # Only fill blanks -- never clobber a manually-entered value.
        updates = {}
        for k, v in fields.items():
            if v in (None, ""):
                continue
            existing = row[k] if k in row.keys() else None
            if existing in (None, ""):
                updates[k] = v
        if updates:
            set_clause = ", ".join(f"{k} = ?" for k in updates)
            conn.execute(
                f"UPDATE accounts SET {set_clause}, updated_at = datetime('now') WHERE id = ?",
                list(updates.values()) + [row["id"]],
            )
        return row["id"]


def get_or_create_brand(conn, name, category=None):
    name = _clean(name)
    if not name:
        return None
    row = conn.execute("SELECT id FROM brands WHERE name = ? COLLATE NOCASE", (name,)).fetchone()
    if row:
        return row["id"]
    admin_cat = "admin" if "documents" in name.lower() or name.lower().startswith("big rivers marketing") else category
    conn.execute("INSERT INTO brands (name, category) VALUES (?, ?)", (name, admin_cat))
    return conn.execute("SELECT id FROM brands WHERE name = ? COLLATE NOCASE", (name,)).fetchone()["id"]


def get_or_create_contact(conn, account_id, name, source="manual"):
    name = _clean(name)
    if not name or not account_id:
        return None
    row = conn.execute(
        "SELECT id FROM contacts WHERE account_id = ? AND name = ? COLLATE NOCASE",
        (account_id, name),
    ).fetchone()
    if row:
        return row["id"]
    conn.execute(
        "INSERT INTO contacts (account_id, name, source) VALUES (?, ?, ?)",
        (account_id, name, source),
    )
    return conn.execute(
        "SELECT id FROM contacts WHERE account_id = ? AND name = ? COLLATE NOCASE",
        (account_id, name),
    ).fetchone()["id"]


def _record_batch(conn, batch_type, filename, summary):
    cur = conn.execute(
        "INSERT INTO import_batches (batch_type, filename, summary_json) VALUES (?, ?, ?)",
        (batch_type, filename, json.dumps(summary)),
    )
    return cur.lastrowid


# ---------------------------------------------------------------------------
# Companies importer (account master list)
# ---------------------------------------------------------------------------

def import_companies(conn, path, filename=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    summary = {"rows_seen": 0, "accounts_created": 0, "accounts_updated": 0, "warning": None}
    if not rows:
        _record_batch(conn, "companies", filename or str(path), summary)
        conn.commit()
        return summary

    header = [(_clean(h)).lower() for h in rows[0]]

    def col(*names):
        for n in names:
            if n in header:
                return header.index(n)
        return None

    idx_name = col("name")
    idx_type = col("company type")
    idx_class = col("class")
    idx_category = col("category")
    idx_street = col("street")
    idx_city = col("city")
    idx_state = col("state")
    idx_zip = col("zip code", "zip")
    idx_phone = col("phone 1", "phone")
    idx_website = col("website")
    idx_status = col("status")

    data_rows = [r for r in rows[1:] if any(v not in (None, "") for v in r)]
    summary["rows_seen"] = len(data_rows)

    if idx_name is None or not data_rows:
        summary["warning"] = (
            "Companies file had no data rows (headers only). Accounts were instead derived "
            "from the Activity Journal and Sales files. Re-import a populated Companies export "
            "any time to fill in addresses, class, category, phone, etc."
        )
        _record_batch(conn, "companies", filename or str(path), summary)
        conn.commit()
        return summary

    for r in data_rows:
        name = _clean(r[idx_name]) if idx_name is not None else ""
        if not name:
            continue
        existing = conn.execute("SELECT id FROM accounts WHERE name = ? COLLATE NOCASE", (name,)).fetchone()
        city = _clean(r[idx_city]) if idx_city is not None else ""
        fields = dict(
            company_type=_clean(r[idx_type]) if idx_type is not None else None,
            **{"class": _clean(r[idx_class]) if idx_class is not None else None},
            category=_clean(r[idx_category]) if idx_category is not None else None,
            street=_clean(r[idx_street]) if idx_street is not None else None,
            city=city or None,
            state=_clean(r[idx_state]) if idx_state is not None else None,
            zip=_clean(r[idx_zip]) if idx_zip is not None else None,
            phone=_clean(r[idx_phone]) if idx_phone is not None else None,
            website=_clean(r[idx_website]) if idx_website is not None else None,
            status=_clean(r[idx_status]) if idx_status is not None else None,
            market=infer_market(name, city),
            source="companies_import",
        )
        upsert_account(conn, name, **fields)
        if existing:
            summary["accounts_updated"] += 1
        else:
            summary["accounts_created"] += 1

    _record_batch(conn, "companies", filename or str(path), summary)
    conn.commit()
    return summary


# ---------------------------------------------------------------------------
# Activity Journal importer (call history)
# ---------------------------------------------------------------------------

def import_journal(conn, path, filename=None, rep_filter="Nathen Bitzer"):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    summary = {
        "rows_seen": 0,
        "calls_imported": 0,
        "calls_skipped_duplicate": 0,
        "accounts_created": 0,
        "contacts_created": 0,
        "followups_created": 0,
        "brands_seen": set(),
    }

    for r in rows:
        r = list(r) + [None] * (11 - len(r))
        sales_rep = _clean(r[10])
        if rep_filter and sales_rep.lower() != rep_filter.lower():
            continue
        summary["rows_seen"] += 1

        call_date = _parse_date(r[0])
        subject = _clean(r[1])
        company = _clean(r[2])
        company_type = _clean(r[3])
        attendees_raw = _clean(r[4])
        report_type = _clean(r[5]) or subject
        brand_name = _clean(r[7])
        notes = _clean(r[8])
        followup_date = _parse_date(r[9])

        if not company or not call_date:
            continue

        before = conn.execute("SELECT COUNT(*) c FROM accounts").fetchone()["c"]
        account_id = upsert_account(
            conn, company, company_type=company_type or None, market=infer_market(company)
        )
        after = conn.execute("SELECT COUNT(*) c FROM accounts").fetchone()["c"]
        if after > before:
            summary["accounts_created"] += 1

        brand_id = None
        if brand_name:
            brand_id = get_or_create_brand(conn, brand_name)
            summary["brands_seen"].add(brand_name)

        attendee_names = _split_attendees(attendees_raw)
        primary_contact_id = None
        for i, aname in enumerate(attendee_names):
            before_c = conn.execute("SELECT COUNT(*) c FROM contacts").fetchone()["c"]
            cid = get_or_create_contact(conn, account_id, aname, source="journal_import")
            after_c = conn.execute("SELECT COUNT(*) c FROM contacts").fetchone()["c"]
            if after_c > before_c:
                summary["contacts_created"] += 1
            if i == 0:
                primary_contact_id = cid

        import_key = hashlib.sha1(
            "|".join([call_date, company.lower(), brand_name.lower(), subject.lower(), notes[:200].lower()]).encode(
                "utf-8"
            )
        ).hexdigest()

        cur = conn.execute(
            """INSERT OR IGNORE INTO calls
               (account_id, contact_id, call_date, report_type, subject, brand_id, attendees, notes,
                followup_date, source, import_key)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 'journal_import', ?)""",
            (
                account_id,
                primary_contact_id,
                call_date,
                report_type,
                subject,
                brand_id,
                attendees_raw,
                notes,
                followup_date,
                import_key,
            ),
        )
        if cur.rowcount == 0:
            summary["calls_skipped_duplicate"] += 1
            continue

        summary["calls_imported"] += 1
        call_id = cur.lastrowid

        if followup_date:
            dup = conn.execute("SELECT id FROM followups WHERE call_id = ?", (call_id,)).fetchone()
            if not dup:
                desc = f"Follow up: {subject or report_type}" + (f" ({brand_name})" if brand_name else "")
                conn.execute(
                    """INSERT INTO followups (account_id, contact_id, call_id, description, due_date, source)
                       VALUES (?, ?, ?, ?, ?, 'auto')""",
                    (account_id, primary_contact_id, call_id, desc, followup_date),
                )
                summary["followups_created"] += 1

    summary["brands_seen"] = sorted(summary["brands_seen"])
    _record_batch(conn, "journal", filename or str(path), summary)
    conn.commit()
    return summary


# ---------------------------------------------------------------------------
# Sales numbers importer (account -> brand -> YTD / prior-YTD / variance)
# ---------------------------------------------------------------------------

def import_sales(conn, path, filename=None):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    summary = {
        "accounts_seen": 0,
        "accounts_created": 0,
        "brand_lines_imported": 0,
        "brands_seen": set(),
        "total_current_ytd": 0.0,
        "total_prior_ytd": 0.0,
    }

    batch_id = _record_batch(conn, "sales", filename or str(path), {})

    state = "seek_header"
    current_account_id = None
    current_current_label = None
    current_prior_label = None

    for r in rows:
        a, b, c, d = (list(r) + [None, None, None, None])[:4]
        a_str = _clean(a)

        if a is None and b is None:
            state = "seek_header"
            continue

        if isinstance(b, str) and "YTD" in b:
            before = conn.execute("SELECT COUNT(*) c FROM accounts").fetchone()["c"]
            current_account_id = upsert_account(conn, a_str, market=infer_market(a_str))
            after = conn.execute("SELECT COUNT(*) c FROM accounts").fetchone()["c"]
            if after > before:
                summary["accounts_created"] += 1
            summary["accounts_seen"] += 1
            current_current_label = b
            current_prior_label = _clean(c)
            state = "in_account"
            continue

        if state == "in_account":
            if a_str.endswith("Total"):
                state = "seek_header"
                continue
            if not a_str:
                continue
            brand_name = a_str
            current_ytd = b if isinstance(b, (int, float)) else 0.0
            prior_ytd = c if isinstance(c, (int, float)) else 0.0
            variance = d if isinstance(d, (int, float)) else (current_ytd - prior_ytd)

            brand_id = get_or_create_brand(conn, brand_name)
            conn.execute(
                """INSERT INTO sales_snapshots
                   (account_id, brand_id, current_label, prior_label, current_ytd, prior_ytd, variance, import_batch_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    current_account_id,
                    brand_id,
                    current_current_label,
                    current_prior_label,
                    current_ytd,
                    prior_ytd,
                    variance,
                    batch_id,
                ),
            )
            summary["brand_lines_imported"] += 1
            summary["brands_seen"].add(brand_name)
            summary["total_current_ytd"] += current_ytd
            summary["total_prior_ytd"] += prior_ytd

    summary["brands_seen"] = sorted(summary["brands_seen"])
    conn.execute(
        "UPDATE import_batches SET summary_json = ? WHERE id = ?",
        (json.dumps(summary), batch_id),
    )
    conn.commit()
    return summary
